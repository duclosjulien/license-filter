import openpyxl
workbook = openpyxl.load_workbook('ProjectJu.xlsx')
sheet = workbook['Project Online Mar 21']

column_index = 11 
filtered_rows = []

# Iterate through all rows
for row in sheet.iter_rows():
    if row[column_index].value is not None:
        # Append the row to the filtered list if the first cell is not None
        filtered_rows.append(row)

# Clear existing rows in the sheet
sheet.delete_rows(1, sheet.max_row)

# Write the filtered rows back to the sheet
for filtered_row in filtered_rows:
   sheet.append([cell.value for cell in filtered_row])

workbook.save('ModifiedList.xlsx')
workbook.close()
